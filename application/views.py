import os
import re
import json
from tempfile import NamedTemporaryFile

from django.conf import settings
from django.http import HttpResponse,JsonResponse
from django.shortcuts import render, redirect

from openpyxl import Workbook
from django.contrib import messages
from application.models import *
from django.views.decorators.csrf import csrf_exempt


def index(request):

    if request.method == "GET":
        return render(request, "application/index.html", {})



@csrf_exempt
def generateExcelSheet(request):

    if request.method == "GET":
        context = {"entries": Entry.objects.all()}
        print(context)
        return render(request, "application/generation_form.html", context)

    if request.method == "POST":
        entry = Entry.objects.get(id=request.POST["entry"])

        wb = Workbook()
        ws = wb.active

        ws["A1"] = "Row #"
        ws["B1"] = "People"

        ws.column_dimensions["B"].width = 200

        for i, row in enumerate(entry.rows.all(), start=2):
            ws["A{}".format(i)] = "Row {}".format(row.level)
            ws["B{}".format(i)] = row.people

        # after generation
        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
            response = HttpResponse(stream, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response['Content-Disposition'] = "attachment; filename={}".format(entry.name)
            return response


@csrf_exempt # ajax
def submit(request):
    if request.method == "POST":
        # passing JSON data would probably be a good idea, in which case:
        """
        sample request:
        {
    	"entry": "RecNAcc",
    	"rows": [
        		{
        			"level": 0,
        			"people": "Hemanth, Nayan, Sanchit, Raghav"
        		},
        		{
        			"level": 1,
        			"people": "Hemanth, Nayan, Sanchit, Raghav"
        		}
    		]
        }
        """
        data = json.loads(request.body.decode('utf-8'))
        try:
            entry = Entry.objects.get_or_create(name=data["entry"])[0]
        except:
            response=JsonResponse({'message':'Department exists'})
            return response
        print(data)
        for row in data["rows"]:
            row_to_add = Row.objects.get_or_create(level=row["level"], entry=entry)[0]
            row_to_add.people = row["people"]
            row_to_add.save()
        entry.save()
        messages.success(request,'Enteries added')
        return JsonResponse({'message':'Done'})


@csrf_exempt
def autocomplete(request): # for ajax
        if request.method == "POST":
            """
            sample request:
                {
                	"snippet": "hem"
                }
            """
            req_data = json.loads(request.body.decode('utf-8'))

            mess_list = list()
            response_data = {"found": []}
            snippet = re.compile(r'^{}'.format(req_data["snippet"].upper()))

            with open(os.path.join(settings.BASE_DIR, "mess_list.json"), "r") as file:
                mess_list = json.loads(file.read().encode('utf-8-sig').decode('utf-8-sig')) # 0_0

            for val in mess_list:
                if snippet.match(val["SNAME"]):
                    response_data["found"].append(val["SNAME"])

            return JsonResponse(response_data)
